#!/usr/bin/env python3
"""Import approved input cells from the investment workbook into the website."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import shutil
import subprocess
import sys
from datetime import datetime
from pathlib import Path

import openpyxl


SYMBOLS = ("VOO", "NVDA", "0050", "0056", "00919", "00631L")
US_SYMBOLS = {"VOO", "NVDA"}
BLOCK_PATTERN = re.compile(
    r"  // BEGIN WORKBOOK IMPORT\r?\n.*?  // END WORKBOOK IMPORT",
    re.DOTALL,
)
VERSION_PATTERN = re.compile(r"const DATA_VERSION = (\d+);")


def number(value, label: str, *, empty_zero: bool = False) -> float:
    if value is None and empty_zero:
        return 0.0
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ValueError(f"{label} 必須是數字，目前為 {value!r}")
    return float(value)


def compact(value: float):
    rounded = round(value)
    return int(rounded) if abs(value - rounded) < 1e-9 else round(value, 8)


def assert_label(sheet, cell: str, expected: str) -> None:
    actual = str(sheet[cell].value or "").replace("\n", " ").strip()
    if expected not in actual:
        raise ValueError(f"{sheet.title}!{cell} 欄位格式不符：預期包含「{expected}」，實際為「{actual}」")


def resolve_workbook(path_text: str | None) -> Path:
    if path_text:
        path = Path(path_text).expanduser().resolve()
        if not path.is_file():
            raise FileNotFoundError(f"找不到 Excel：{path}")
        return path

    candidates = list(Path.cwd().glob("投資試算表*.xlsx"))
    candidates += list((Path.home() / "Desktop").glob("投資試算表*.xlsx"))
    if not candidates:
        raise FileNotFoundError("找不到投資試算表 Excel，請提供檔案路徑。")
    return max(candidates, key=lambda item: item.stat().st_mtime)


def parse_plan_cell(formula_value, cached_value, symbol: str, fx_rate: float, label: str):
    if symbol in US_SYMBOLS:
        if isinstance(formula_value, str):
            match = re.match(r"^=\s*([0-9.]+)\s*\*", formula_value)
            if match:
                return compact(float(match.group(1)))
        return compact(number(cached_value, label) / fx_rate)
    return compact(number(cached_value, label, empty_zero=True))


def read_workbook(path: Path, quote_date: str | None) -> dict:
    formulas = openpyxl.load_workbook(path, data_only=False, read_only=True)
    values = openpyxl.load_workbook(path, data_only=True, read_only=True)
    required = {"設定", "持股現況", "投資計畫"}
    missing = required.difference(formulas.sheetnames)
    if missing:
        raise ValueError(f"缺少必要分頁：{', '.join(sorted(missing))}")

    settings_f = formulas["設定"]
    settings_v = values["設定"]
    holdings_f = formulas["持股現況"]
    holdings_v = values["持股現況"]
    plans_f = formulas["投資計畫"]
    plans_v = values["投資計畫"]

    assert_label(settings_f, "B5", "起始年度")
    assert_label(settings_f, "B11", "美元匯率")
    assert_label(settings_f, "B28", "銀行期初－台幣")
    assert_label(settings_f, "B34", "購車年份")
    assert_label(holdings_f, "B4", "代號")
    assert_label(holdings_f, "H4", "期初持有")
    assert_label(plans_f, "B4", "西元年")
    assert_label(plans_f, "E4", "VOO")

    def setting(cell: str, label: str, *, empty_zero: bool = False):
        return number(settings_v[cell].value, f"設定!{cell} {label}", empty_zero=empty_zero)

    start_year = int(setting("C5", "起始年度"))
    start_age = int(setting("C6", "起始年齡"))
    end_age = int(setting("C7", "試算到幾歲"))
    first_year_months = int(setting("C24", "首年計入月數"))
    if not 0 <= first_year_months <= 12:
        raise ValueError("設定!C24 首年計入月數必須介於 0 到 12。")
    first_year_data_month = 13 - first_year_months if first_year_months else 12
    fx_rate = setting("C11", "美元匯率")

    source_modified = formulas.properties.modified
    if source_modified is None:
        source_modified = datetime.fromtimestamp(path.stat().st_mtime)
    source_date = quote_date or source_modified.date().isoformat()
    datetime.fromisoformat(source_date)

    imported_settings = {
        "startYear": start_year,
        "startAge": start_age,
        "endAge": end_age,
        "autoRollFirstYearMonths": True,
        "firstYearMonths": first_year_months,
        "firstYearDataMonth": first_year_data_month,
        "fxRate": compact(fx_rate),
        "twPriceGrowth": compact(setting("C12", "台股股價年成長率") * 100),
        "usPriceGrowth": compact(setting("C13", "美股股價年成長率") * 100),
        "twDividendGrowth": compact(setting("C14", "台股配息年成長率") * 100),
        "usDividendGrowth": compact(setting("C15", "美股配息年成長率") * 100),
        "monthlySalary": compact(setting("C18", "月薪")),
        "salaryGrowth": compact(setting("C19", "薪資年成長率") * 100),
        "familyMonthly": compact(setting("C20", "孝親")),
        "leisureMonthly": compact(setting("C21", "特別預算")),
        "fixedMonthly": compact(setting("C22", "固定開銷")),
        "expenseInflation": compact(setting("C23", "支出年成長率") * 100),
        "annualSalaryMonths": compact(setting("C25", "年薪月數")),
        "twdDeposit": compact(setting("C28", "台幣存款")),
        "foreignDepositTwd": compact(setting("C29", "外幣存款")),
        "bankMinimum": compact(setting("C31", "銀行存款年底下限")),
        "carYear": compact(setting("C34", "購車年份")),
        "carPrice": compact(setting("C35", "購車總價")),
        "carDownPaymentRate": compact(setting("C36", "購車頭期比例") * 100),
        "carLoanRate": compact(setting("C37", "車貸年利率") * 100),
        "carLoanMonths": compact(setting("C38", "車貸期數")),
        "annualVehicleCost": compact(setting("C39", "車輛年成本")),
        "homeYear": compact(setting("C47", "購屋年份", empty_zero=True)),
        "homePrice": compact(setting("C48", "房屋總價", empty_zero=True)),
        "homeDownPaymentRate": compact(setting("C49", "購屋頭期比例", empty_zero=True) * 100),
        "homeLoanRate": compact(setting("C50", "房貸年利率", empty_zero=True) * 100),
        "homeLoanMonths": compact(setting("C51", "房貸期數", empty_zero=True)),
        "annualHomeCost": compact(setting("C52", "房屋年成本", empty_zero=True)),
        "updatedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
        "quoteStatus": "Excel 附檔匯入（配息已折減 20%）",
        "quoteDates": {"TW": source_date, "US": source_date, "FX": source_date},
    }

    row_by_symbol = {}
    for row in range(5, holdings_f.max_row + 1):
        symbol = str(holdings_f.cell(row, 2).value or "").strip()
        if symbol in SYMBOLS:
            row_by_symbol[symbol] = row
    missing_symbols = set(SYMBOLS).difference(row_by_symbol)
    if missing_symbols:
        raise ValueError(f"持股現況缺少標的：{', '.join(sorted(missing_symbols))}")

    holding_settings = {}
    payout_months = {}
    for symbol in SYMBOLS:
        row = row_by_symbol[symbol]
        holding_settings[symbol] = {
            "units": compact(number(holdings_v.cell(row, 8).value, f"持股現況!H{row}", empty_zero=True)),
            "cost": compact(number(holdings_v.cell(row, 6).value, f"持股現況!F{row}", empty_zero=True)),
            "price": compact(number(holdings_v.cell(row, 7).value, f"持股現況!G{row}", empty_zero=True)),
            "annualDividend": compact(number(holdings_v.cell(row, 9).value, f"持股現況!I{row}", empty_zero=True)),
        }
        payout_months[symbol] = [
            month for month in range(1, 13)
            if bool(holdings_v.cell(row, 17 + month).value)
        ]

    final_year = start_year + end_age - start_age
    manual_plans = {}
    for row in range(5, plans_f.max_row + 1):
        year_value = plans_v.cell(row, 2).value
        if not isinstance(year_value, (int, float)):
            continue
        year = int(year_value)
        if not start_year <= year <= final_year:
            continue
        manual_plans[str(year)] = {}
        for offset, symbol in enumerate(SYMBOLS, start=5):
            manual_plans[str(year)][symbol] = parse_plan_cell(
                plans_f.cell(row, offset).value,
                plans_v.cell(row, offset).value,
                symbol,
                fx_rate,
                f"投資計畫!{plans_v.cell(row, offset).coordinate}",
            )
    missing_years = [year for year in range(start_year, final_year + 1) if str(year) not in manual_plans]
    if missing_years:
        raise ValueError(f"投資計畫缺少年度：{', '.join(map(str, missing_years))}")

    return {
        "sourceFile": path.name,
        "sourceModifiedAt": source_modified.isoformat(timespec="seconds"),
        "sourceSha256": hashlib.sha256(path.read_bytes()).hexdigest(),
        "settings": imported_settings,
        "holdingSettings": holding_settings,
        "payoutMonths": payout_months,
        "manualPlans": manual_plans,
    }


def generated_block(data: dict) -> str:
    payload = json.dumps(data, ensure_ascii=False, indent=2)
    indented = "\n".join(f"  {line}" for line in payload.splitlines())
    return f"  // BEGIN WORKBOOK IMPORT\n  const WORKBOOK_DATA = Object.freeze({indented.lstrip()});\n  // END WORKBOOK IMPORT"


def rebuild_deploy(root: Path) -> None:
    deploy = root / "deploy"
    deploy.mkdir(exist_ok=True)
    desktop = (root / "web" / "index.html").read_text(encoding="utf-8")
    mobile = (root / "mobile" / "index.html").read_text(encoding="utf-8")
    (deploy / "index.html").write_text(desktop.replace("../shared/", ""), encoding="utf-8")
    (deploy / "mobile.html").write_text(
        mobile.replace("../web/styles.css", "styles.css").replace("../shared/", ""),
        encoding="utf-8",
    )
    for source, target in (
        (root / "web" / "styles.css", deploy / "styles.css"),
        (root / "shared" / "core.js", deploy / "core.js"),
        (root / "shared" / "charts.js", deploy / "charts.js"),
        (root / "shared" / "app.js", deploy / "app.js"),
        (root / "shared" / "d3.min.js", deploy / "d3.min.js"),
        (root / "README.md", deploy / "README.md"),
    ):
        shutil.copyfile(source, target)


def apply_import(root: Path, data: dict) -> int:
    core_path = root / "shared" / "core.js"
    core_text = core_path.read_text(encoding="utf-8")
    version_match = VERSION_PATTERN.search(core_text)
    if not version_match:
        raise ValueError("shared/core.js 找不到 DATA_VERSION。")
    current_version = int(version_match.group(1))
    new_version = current_version + 1
    updated = BLOCK_PATTERN.sub(generated_block(data), core_text, count=1)
    if updated == core_text:
        raise ValueError("shared/core.js 找不到 WORKBOOK IMPORT 區塊。")
    updated = VERSION_PATTERN.sub(f"const DATA_VERSION = {new_version};", updated, count=1)
    core_path.write_text(updated, encoding="utf-8")

    for relative in ("web/index.html", "mobile/index.html"):
        path = root / relative
        html = path.read_text(encoding="utf-8")
        html = re.sub(r"((?:styles|core|app)\.js|styles\.css)\?v=\d+", rf"\1?v={new_version}", html)
        path.write_text(html, encoding="utf-8")

    rebuild_deploy(root)
    return new_version


def main() -> int:
    parser = argparse.ArgumentParser(description="將投資 Excel 的核准輸入欄位匯入桌機版與手機版網站。")
    parser.add_argument("workbook", nargs="?", help="Excel 路徑；省略時尋找最新的投資試算表*.xlsx")
    parser.add_argument("--quote-date", help="附檔市場資料基準日，格式 YYYY-MM-DD；預設使用 Excel 修改日")
    parser.add_argument("--apply", action="store_true", help="通過驗證後實際更新網站；未指定時只顯示預覽")
    parser.add_argument("--publish", action="store_true", help="更新後呼叫 GitHub Pages 發布程式")
    args = parser.parse_args()
    if args.publish and not args.apply:
        parser.error("--publish 必須搭配 --apply")

    root = Path(__file__).resolve().parents[1]
    workbook = resolve_workbook(args.workbook)
    data = read_workbook(workbook, args.quote_date)
    summary = {
        "workbook": str(workbook),
        "sourceModifiedAt": data["sourceModifiedAt"],
        "planning": {
            "startYear": data["settings"]["startYear"],
            "endAge": data["settings"]["endAge"],
            "firstYearMonths": data["settings"]["firstYearMonths"],
        },
        "cash": {
            "twd": data["settings"]["twdDeposit"],
            "foreignTwd": data["settings"]["foreignDepositTwd"],
        },
        "holdings": data["holdingSettings"],
        "planYears": len(data["manualPlans"]),
        "mode": "apply" if args.apply else "preview",
    }

    if args.apply:
        summary["dataVersion"] = apply_import(root, data)
        if args.publish:
            command = [
                "powershell.exe", "-NoProfile", "-ExecutionPolicy", "Bypass",
                "-File", str(root / "scripts" / "publish-github-pages.ps1"),
                "-Message", f"Import {workbook.name}",
            ]
            subprocess.run(command, cwd=root, check=True)
            summary["published"] = True

    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as error:
        print(f"匯入失敗：{error}", file=sys.stderr)
        raise SystemExit(1)
